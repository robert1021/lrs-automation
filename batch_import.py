import openpyxl
import pandas as pd
import os
from datetime import datetime
from constants import *


class BatchImport:
    """ This class is used to generate the LRS batch import report.

        * You must use save_file() after you are finished working with the Excel file to save

    """

    def __init__(self):
        self.batch_import_path = None

    def create_lrs_batch_import_report(self):
        report_name = f"Batch Imports {datetime.today().strftime('%Y-%m-%d')}.xlsx"
        self.batch_import_path = os.path.join("Output", report_name)
        report = {
            "PREFIX": [],
            "FILE NUMBER": [],
            "FILE SECURITY": [],
            "FILE TYPE": [],
            "ESSENTIAL": [],
            "STORAGE": [],
            "DATE FILE CREATED": [],
            "FILE STATUS": [],
            "FILE STATUS DATE": [],
            "ENGLISH FILE TITLE": [],
            "ENGLISH DESCRIPTION": [],
            "FRENCH FILE TITLE": [],
            "FRENCH DESCRIPTION": [],
            "ORANIZATIONAL CODE": [],
            "FOLDER PREFIX": [],
            "FOLDER FILE NUMBER": [],
            "FOLDER FILE SECURITY": [],
            "FOLDER TYPE": [],
            "FOLDER STATUS": [],
            "FOLDER SECURITY": [],
            "LOCATION": [],
            "FOLDER FROM DATE": [],
            "FOLDER TO DATE": [],
            "COMMENT": [],
            "CONTENT": [],
            "REFERENCE": [],
            "REFERENCE TYPE": []
        }
        df = pd.DataFrame(report)
        df.to_excel(self.batch_import_path, index=False)

    def add_to_report(self, data_to_add_df: pd.DataFrame):

        # create workbook object
        wb_obj = openpyxl.load_workbook(self.batch_import_path)
        ws = wb_obj['Sheet1']

        row_count = len([row for row in wb_obj['Sheet1'] if not all([cell.value is None for cell in row])])

        for row in data_to_add_df.itertuples():
            row_count += 1

            # print(row)

            # Determine file and folder status

            # Withdrawn, Refused, cancelled, discontinued, rejected
            if str(row.status_licence).lower() in ['refused', 'withdrawn', 'cancelled', 'discontinued', 'rejected']:
                status = 'DOR'

            # All other
            else:
                status = 'ACT'

            ws[f'{PREFIX}{row_count}'] = 'F-FCS'
            ws[f'{FILE_NUMBER}{row_count}'] = str(row.lrs_file_number)
            ws[f'{FILE_SECURITY}{row_count}'] = 'PRO'
            ws[f'{FILE_TYPE}{row_count}'] = 'CAS'
            ws[f'{ESSENTIAL}{row_count}'] = 'NES'
            ws[f'{STORAGE}{row_count}'] = 'DI'
            ws[f'{DATE_FILE_CREATED}{row_count}'] = str(datetime.today().strftime('%Y-%m-%d'))
            ws[f'{FILE_STATUS}{row_count}'] = status
            ws[f'{FILE_STATUS_DATE}{row_count}'] = str(datetime.today().strftime('%Y-%m-%d'))
            ws[f'{ENGLISH_FILE_TITLE}{row_count}'] = str(
                row.product_name) if 'product_name' in data_to_add_df.columns else str(row.company_name)
            ws[f'{FRENCH_FILE_TITLE}{row_count}'] = str(
                row.product_name) if 'product_name' in data_to_add_df.columns else str(row.company_name)
            ws[f'{ORGANIZATIONAL_CODE}{row_count}'] = 'F1G'
            ws[f'{FOLDER_PREFIX}{row_count}'] = 'F-FCS'
            ws[f'{FOLDER_FILE_NUMBER}{row_count}'] = str(row.lrs_file_number)
            ws[f'{FOLDER_FILE_SECURITY}{row_count}'] = 'PRO'
            ws[f'{FOLDER_TYPE}{row_count}'] = 'DB'
            ws[f'{FOLDER_STATUS}{row_count}'] = status
            ws[f'{FOLDER_SECURITY}{row_count}'] = 'PRO'
            ws[f'{LOCATION}{row_count}'] = 'X02'
            ws[f'{FOLDER_FROM_DATE}{row_count}'] = str(row.received_date).split(' ')[0]
            ws[f'{FOLDER_TO_DATE}{row_count}'] = str(row.completed_date).split(' ')[0] if status == 'DOR' else None

            # Determine the comment

            if str(row.licence_number) == 'nan':
                comment = row.status_licence
            else:
                comment = f'{row.status_licence} - {int(row.licence_number)}'

            ws[f'{COMMENT}{row_count}'] = comment

        # save after rows added
        wb_obj.save(self.batch_import_path)
